import openpyxl
import requests
import json
import csv

workbook = openpyxl.load_workbook(r'C:\Users\Admin\PycharmProjects\caodata\pluscode.xlsx')
worksheet = workbook['Sheet1']

row = 2  # Start from row 2 (A2)
cell_value = worksheet.cell(row=row, column=1).value
while cell_value is not None:
    url = 'https://api.map4d.vn/sdk/v2/geocode'
    params = {
        'address': f'{cell_value}',
        'key': '93d393d0f6507ee00b62fe01db7430fa'
    }
    headers = {
        'Accept': 'text/plain'
    }
    response = requests.get(url, params=params, headers=headers)
    data = response.json()
    print(data)
    results = data['result']
    with open('ancu.csv', 'a', newline='', encoding='utf-8') as file_output:
        headers = ['pluscode', 'location']
        writer = csv.DictWriter(file_output, delimiter=',', lineterminator='\n', fieldnames=headers)
        if row == 2:
            writer.writeheader()
        if len(results) > 0:
            first_result_location = results[0]['location']
            print(cell_value)
            writer.writerow({headers[0]: cell_value, headers[1]: first_result_location})
    row += 1
    cell_value = worksheet.cell(row=row, column=1).value
