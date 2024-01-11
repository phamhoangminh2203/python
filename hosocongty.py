import pyautogui
from selenium import webdriver
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
import csv
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import NoSuchElementException
import openpyxl
excel_file_path = r'C:\Users\Admin\PycharmProjects\caodata\foody.xlsx'

# Initialize the WebDriver and open the Excel file
edge_driver_path = "C:/Users/Admin/PycharmProjects/caodata/venv/msedgedriver.exe"
service_obj = Service(edge_driver_path)
browser = webdriver.Edge(service=service_obj)
browser.maximize_window()

workbook = openpyxl.load_workbook(excel_file_path)
worksheet = workbook['Sheet1']
row = 2  # Start from row 2 (assuming row 1 is a header)
cell_value = worksheet.cell(row=row, column=1).value

with open('danhnghiep.csv', 'w', newline='', encoding='utf-8') as file_output:
    headers = ['Name', 'mast', 'diachi', 'daidien', 'ngay', 'tinhtrang']
    writer = csv.DictWriter(file_output, delimiter=',', lineterminator='\n', fieldnames=headers)
    writer.writeheader()
    def GetURL():
        page_source = BeautifulSoup(browser.page_source, 'html.parser')
        profiles = page_source.find_all('ul', class_='hsdn')
        all_profile_URL = []
        for profile in profiles:
            # Find the 'a' element within the 'h3' element and extract the 'href' attribute
            anchor = profile.find('a')
            if anchor:
                a="https://hosocongty.vn/"
                link = anchor.get('href')
                profile_ID = a+link
                if profile_ID not in all_profile_URL:
                    all_profile_URL.append(profile_ID)
        return all_profile_URL
    browser.get(cell_value)
    URLs_all_page = GetURL()
    URLs_all_page = [page for page in GetURL()]
    print(URLs_all_page)
    for Googlemap_URL in URLs_all_page:
        browser.get(Googlemap_URL)
        page_source = BeautifulSoup(browser.page_source, 'html.parser')
        info_section = page_source.find('div', class_='box_content')
        company_name = info_section.find('h1').text.strip()
        # Search for relevant information by label or icon
        info_items = info_section.find_all('li')
        tax_code = address = legal_representative = date_of_registration = status = None
        for item in info_items:
            label = item.find('label')
            if label:
                label_text = label.text.strip()
                if 'Mã số thuế:' in label_text:
                    tax_code = item.find('span').text.strip()
                elif 'Địa chỉ thuế:' in label_text:
                    address = item.find('span').text.strip()
                elif 'Đại diện pháp luật:' in label_text:
                    legal_representative = item.find('a').text.strip()
                elif 'Ngày cấp:' in label_text:
                    date_of_registration = item.find('a').text.strip()
                elif 'Trạng thái:' in label_text:
                    status = item.find('span').text.strip()
            writer.writerow({
                headers[0]: company_name,  # Name
                headers[1]: tax_code,  # Mã số thuế
                headers[2]: address,  # Địa chỉ
                headers[3]: legal_representative,  # Đại diện pháp luật
                headers[4]: date_of_registration,  # Ngày cấp
                headers[5]: status,  # Trạng thái
            })
row += 1
cell_value = worksheet.cell(row=row, column=1).value  # Update cell_value
browser.quit()